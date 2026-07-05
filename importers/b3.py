"""Minimal B3 XLSX importer.

This importer is intentionally small and explicit:

- identify files by filename
- assume the XLSX table layout is stable
- raise on parsing errors
- log each row decision
"""

from __future__ import annotations

import logging
import os
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Optional

import beangulp
import openpyxl
from beancount.core import amount, data
from beancount.core.position import Cost, CostSpec
from beancount.core.number import MISSING
from beancount.parser.grammar import ValueType


logger = logging.getLogger(__name__)

BROKERS = {
    "XP INVESTIMENTOS CCTVM S/A": "XP",
    "XP SERVICOS FINANCEIROS DTVM LTDA": "XP",
    "XP INVESTIMENTOS CORRETORA DE CAMBIO, TITULOS E VALORES MOBI": "XP",
    "XP INVESTIMENTOS CORRETORA DE CAMBIO TITULOS E VALORES MOBILIARIOS S/A": "XP",
    "MODAL DTVM LTDA": "XP",
    "CLEAR CORRETORA - GRUPO XP": "XP",
    "INTER DISTRIBUIDORA DE TITULOS E VALORES MOBILIARIOS LTDA": "Inter",
    "BANCO DO BRASIL S/A": "BB",
}

IGNORED_MOVEMENTS = {
    "ATUALIZACAO",
    "DIREITO DE SUBSCRICAO",
    "DIREITOS DE SUBSCRICAO - EXERCIDO",
    "DIREITOS DE SUBSCRICAO - NAO EXERCIDO",
    "CESSAO DE DIREITOS",
    "CESSAO DE DIREITOS - SOLICITADA",
    "RECIBO DE SUBSCRICAO",
    "SOLICITACAO DE SUBSCRICAO",
    "TRANSFERENCIA - LIQUIDACAO",
    "RESGATE",
    "CISAO",
    "INCORPORACAO",
    "FRACAO EM ATIVOS",
}


def _strip_accents(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = _strip_accents(str(value)).upper().strip()
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def parse_date(value: Any, *, context: str) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip() if value is not None else ""
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{context}: invalid date value {value!r}")


def parse_decimal(value: Any, *, context: str) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text or text == "-":
        return None

    text = text.replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)

    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{context}: invalid decimal value {value!r}") from exc


def normalize_ticker(product: Any, movement: str) -> str:
    text = str(product).strip()
    if " - " in text:
        candidate = text.split(" - ", 1)[0].strip()
        candidate_norm = normalize_text(candidate)
        if candidate_norm and candidate_norm.replace(" ", "").isalnum() and any(ch.isdigit() for ch in candidate_norm) and len(candidate_norm) <= 6:
            text = candidate

    normalized_text = normalize_text(text)
    if re.match(r"^[A-Z]{4}\d{1,4}F?$", normalized_text):
        return normalized_text[:-1] if normalized_text.endswith("F") else normalized_text

    # Some B3 products (for example Tesouro Direto) are imported with human
    # readable names that include spaces or symbols like '+'. Convert those
    # names into deterministic Beancount-safe symbols that are valid both as
    # account components and commodity names.
    text = re.sub(r"[^A-Z0-9]+", "-", normalized_text).strip("-")
    text = re.sub(r"-+", "-", text)
    if text and text[0].isdigit():
        text = f"T-{text}"
    return text


def broker_code(institution: Any, *, context: str) -> str:
    normalized = normalize_text(institution)
    if not normalized:
        raise ValueError(f"{context}: missing institution")
    for key, code in BROKERS.items():
        if key in normalized:
            return code
    raise ValueError(f"{context}: unknown institution {institution!r}")


def row_has_content(row: Iterable[Any]) -> bool:
    return any(cell is not None and str(cell).strip() != "" for cell in row)


class B3Importer(beangulp.Importer):
    def __init__(self, account_root: str = "Assets:Investment"):
        self.account_root = account_root.rstrip(":")
        # Cache loaded workbooks by filepath so each file is parsed only once
        # across identify / date / filename / extract calls.
        self._workbook_cache: dict[str, Any] = {}

    def _load_workbook(self, filepath: str):
        if filepath not in self._workbook_cache:
            # B3 xlsx files use features (e.g. shared strings, certain cell
            # styles) that are incompatible with openpyxl's streaming mode.
            # read_only=True causes the workbook to appear as a single cell,
            # so we load the full DOM with read_only=False.
            self._workbook_cache[filepath] = openpyxl.load_workbook(
                filepath, read_only=False, data_only=True
            )
        return self._workbook_cache[filepath]

    # FIX #9: strip accents from the filename before matching, so that files
    # named "movimentação" (with accent) are correctly identified alongside
    # "movimentacao".
    def identify(self, filepath: str) -> bool:
        base = _strip_accents(os.path.basename(filepath)).lower()
        return base.endswith(".xlsx") and ("movimentacao" in base or "negociacao" in base)

    @property
    def name(self) -> str:
        return "B3 Importer"

    def account(self, filepath: str) -> str:
        return f"{self.account_root}:B3"

    # FIX #4: use _load_workbook so the file is not re-opened here if it was
    # already loaded by a prior call (e.g. filename → date → extract).
    def date(self, filepath: str) -> Optional[date]:
        workbook = self._load_workbook(filepath)
        for sheet_name in workbook.sheetnames:
            sheet_key = normalize_text(sheet_name)
            if "NEGOCIACAO" in sheet_key:
                for row in workbook[sheet_name].iter_rows(min_row=2, max_row=10, values_only=True):
                    if row_has_content(row):
                        return parse_date(row[0], context=f"{filepath} / {sheet_name} / date")
            if "MOVIMENTACAO" in sheet_key:
                for row in workbook[sheet_name].iter_rows(min_row=2, max_row=10, values_only=True):
                    if row_has_content(row):
                        return parse_date(row[1], context=f"{filepath} / {sheet_name} / date")
        return None

    def filename(self, filepath: str) -> str:
        base = os.path.basename(filepath)
        file_date = self.date(filepath)
        return f"{file_date:%Y-%m-%d}.{base}" if file_date else base

    # FIX #4 (continued): extract also uses _load_workbook; the workbook was
    # already parsed during date/filename, so no disk I/O happens here.
    def extract(self, filepath: str, existing: list[data.Directive]) -> list[data.Directive]:
        logger.info("Starting B3 import: %s", filepath)
        workbook = self._load_workbook(filepath)
        try:
            entries: list[data.Directive] = []
            for sheet_name in workbook.sheetnames:
                sheet_key = normalize_text(sheet_name)
                if "NEGOCIACAO" in sheet_key:
                    entries.extend(self._extract_negociacoes(filepath, workbook[sheet_name], sheet_name))
                elif "MOVIMENTACAO" in sheet_key:
                    entries.extend(self._extract_movimentacoes(filepath, workbook[sheet_name], sheet_name))

            entries = self._merge_transfers(entries)
            entries.sort(key=lambda entry: (entry.date, entry.meta.get("lineno", 0)))
            logger.info("Finished B3 import: %s (%d entries)", filepath, len(entries))
            return entries
        except Exception:
            logger.exception("Failed importing %s", filepath)
            raise

    def _merge_transfers(self, entries: list[data.Directive]) -> list[data.Directive]:
        """Merge paired custody transfer transactions into a single direct transaction.

        The importer emits each TRANSFERENCIA row as a stub with Equity:Transfers
        as the counterpart. When the same file contains both sides of the transfer
        (same date, same ticker, same absolute quantity, opposite direction), we
        replace the two stubs with one transaction that moves directly from the
        outgoing account to the incoming account. Beancount then carries the cost
        basis across automatically under FIFO.

        Any stub without a matching pair (e.g. transfer to an external broker not
        present in the exports) is left unchanged.
        """
        # Separate transfer stubs from everything else.
        # A transfer stub is identified by the Equity:Transfers posting.
        transfer_stubs: list[data.Transaction] = []
        other_entries: list[data.Directive] = []

        for entry in entries:
            if (
                isinstance(entry, data.Transaction)
                and entry.meta.get("warning") == "unmatched_transfer"
                and any(p.account == "Equity:Transfers" for p in entry.postings)
            ):
                transfer_stubs.append(entry)
            else:
                other_entries.append(entry)

        # Build a pairing key: (date, ticker, abs_quantity).
        # For each key, collect the stubs. Outgoing stubs have a negative
        # quantity on the asset account; incoming stubs have a positive quantity.
            by_key: dict[tuple, list[data.Transaction]] = defaultdict(list)

        for stub in transfer_stubs:
            # The asset posting is the one that is NOT Equity:Transfers.
            asset_posting = next(p for p in stub.postings if p.account != "Equity:Transfers")
            ticker = asset_posting.units.currency
            qty = abs(asset_posting.units.number)
            key = (stub.date, ticker, qty)
            by_key[key].append(stub)

        merged: list[data.Directive] = []

        for key, stubs in by_key.items():
            # Partition into outgoing (negative asset qty) and incoming (positive).
            outgoing = [s for s in stubs if next(p for p in s.postings if p.account != "Equity:Transfers").units.number < 0]
            incoming = [s for s in stubs if next(p for p in s.postings if p.account != "Equity:Transfers").units.number > 0]

            # Pair them off one-to-one in order.
            while outgoing and incoming:
                out_stub = outgoing.pop(0)
                in_stub = incoming.pop(0)

                out_posting = next(p for p in out_stub.postings if p.account != "Equity:Transfers")
                in_posting = next(p for p in in_stub.postings if p.account != "Equity:Transfers")

                # Build merged transaction: outgoing account → incoming account.
                # The outgoing side keeps {} so FIFO can match the lot.
                # The incoming side also uses {} so Beancount assigns the matched cost.
                merged_postings = [
                    data.Posting(out_posting.account, out_posting.units, self._empty_costspec(), None, None, None),
                    data.Posting(in_posting.account, in_posting.units, self._empty_costspec(), None, None, None),
                ]

                # Combine metadata from both stubs; use out_stub as the base.
                meta = dict(out_stub.meta)
                meta["id_in"] = in_stub.meta.get("id", "")
                meta.pop("warning", None)  # no longer unmatched

                txn = data.Transaction(
                    meta,
                    out_stub.date,
                    "*",
                    None,
                    out_stub.narration,
                    frozenset(),
                    frozenset(),
                    merged_postings,
                )
                logger.info(
                    "Merged transfer pair: %s → %s (%s %s)",
                    out_posting.account, in_posting.account, key[2], key[1],
                )
                merged.append(txn)

            # Any leftovers had no match — keep as unmatched stubs.
            for stub in outgoing + incoming:
                logger.warning(
                    "Unmatched transfer stub kept as-is: %s %s qty=%s",
                    stub.date, key[1], key[2],
                )
                merged.append(stub)

        return other_entries + merged

    def _meta(self, filepath: str, lineno: int, **extra: str) -> data.Meta:
        meta = data.new_metadata(filepath, lineno, extra or None)
        return meta

    def _txn(
        self,
        filepath: str,
        lineno: int,
        txn_date: date,
        narration: str,
        postings: list[data.Posting],
        **meta: str,
    ) -> data.Transaction:
        meta_dict = self._meta(filepath, lineno, **meta)
        return data.Transaction(meta_dict, txn_date, "*", None, narration, frozenset(), frozenset(), postings)

    def _amount(self, value: Decimal, currency: str) -> amount.Amount:
        return amount.Amount(value, currency)

    def _cost(self, value: Decimal, txn_date: date) -> Cost:
        return Cost(value, "BRL", txn_date, None)

    def _custom_value(self, value: Any) -> ValueType:
        return ValueType(value, type(value))

    def _empty_costspec(self) -> CostSpec:
        # Emit the parser-equivalent of `{}` so Beancount can book the lot
        # under the configured booking method (AVERAGE).
        return CostSpec(MISSING, None, MISSING, None, None, False)  # type: ignore[arg-type]

    def _asset_account(self, broker: str, ticker: str) -> str:
        return f"{self.account_root}:{broker}:{ticker}"

    def _cash_account(self, broker: str) -> str:
        return f"{self.account_root}:{broker}:Cash"

    def _income_account(self, broker: str, category: str) -> str:
        return f"Income:Investment:{broker}:{category}"

    def _expense_account(self, broker: str, category: str) -> str:
        return f"Expenses:Investment:{broker}:{category}"

    def _extract_negociacoes(self, filepath: str, sheet, sheet_name: str) -> list[data.Directive]:
        entries: list[data.Directive] = []
        for lineno, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_content(row):
                continue

            context = f"{filepath} / {sheet_name} / row {lineno}"
            txn_date = parse_date(row[0], context=context)
            trade_type = normalize_text(row[1])
            institution = broker_code(row[4], context=context)
            ticker = normalize_ticker(row[5], trade_type)
            quantity = parse_decimal(row[6], context=context)
            price = parse_decimal(row[7], context=context)
            value = parse_decimal(row[8], context=context)

            if quantity is None or value is None:
                raise ValueError(f"{context}: missing quantity/value")

            if price is None:
                price = (abs(value) / abs(quantity)) if quantity != 0 else Decimal("0")

            meta = {
                "id": f"b3-neg-{txn_date:%Y%m%d}-{lineno}",
                "source": "b3_negociacoes",
                "asset": ticker,
            }

            cash_account = self._cash_account(institution)
            asset_account = self._asset_account(institution, ticker)

            if trade_type == "COMPRA":
                logger.info("Imported purchase: %s", context)
                gross_value = abs(quantity) * price
                fee_value = abs(value) - gross_value
                postings = [
                    data.Posting(asset_account, self._amount(quantity, ticker), self._cost(price, txn_date), None, None, None),
                    data.Posting(cash_account, self._amount(-abs(value), "BRL"), None, None, None, None),
                ]
                if fee_value != 0:
                    postings.append(
                        data.Posting(
                            self._expense_account(institution, "Fees"),
                            self._amount(fee_value, "BRL"),
                            None, None, None, None,
                        )
                    )
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Buy", postings, **meta))
                continue

            if trade_type == "VENDA":
                logger.info("Imported sale: %s", context)
                gross_value = abs(quantity) * price
                fee_value = gross_value - abs(value)
                postings = [
                    data.Posting(cash_account, self._amount(abs(value), "BRL"), None, None, None, None),
                    data.Posting(asset_account, self._amount(-abs(quantity), ticker), self._empty_costspec(), self._amount(price, "BRL"), None, None),
                ]
                if fee_value != 0:
                    postings.append(
                        data.Posting(
                            self._expense_account(institution, "Fees"),
                            self._amount(fee_value, "BRL"),
                            None, None, None, None,
                        )
                    )
                # Gains leg: amount left blank so Beancount infers it from the
                # difference between cash received and cost basis of the lot(s).
                # Works with FIFO/LIFO. When AVERAGE booking is re-enabled in a
                # future Beancount version, this leg will still be correct.
                postings.append(
                    data.Posting(self._income_account(institution, "Gains"), None, None, None, None, None)
                )
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Sell", postings, **meta))
                continue

            raise ValueError(f"{context}: unsupported negociacao type {trade_type!r}")

        return entries

    def _extract_movimentacoes(self, filepath: str, sheet, sheet_name: str) -> list[data.Directive]:
        entries: list[data.Directive] = []
        for lineno, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row_has_content(row):
                continue

            context = f"{filepath} / {sheet_name} / row {lineno}"
            direction = normalize_text(row[0])
            txn_date = parse_date(row[1], context=context)
            movement = normalize_text(row[2])
            product = str(row[3]).strip() if row[3] is not None else ""
            institution = broker_code(row[4], context=context)
            quantity = parse_decimal(row[5], context=context)
            unit_price = parse_decimal(row[6], context=context)
            value = parse_decimal(row[7], context=context)

            if not movement:
                raise ValueError(f"{context}: missing movement type")

            ticker = normalize_ticker(product, movement)
            cash_account = self._cash_account(institution)
            asset_account = self._asset_account(institution, ticker)

            meta = {
                "id": f"b3-mov-{txn_date:%Y%m%d}-{lineno}",
                "source": "b3_movimentacoes",
                "asset": ticker,
            }

            if movement in IGNORED_MOVEMENTS:
                logger.info("Ignored by rule: %s", context)
                continue

            if movement in {"RENDIMENTO", "DIVIDENDO", "DIVIDENDO - TRANSFERIDO", "RENDIMENTO - TRANSFERIDO"}:
                if value is None:
                    raise ValueError(f"{context}: missing value for income event")
                if "TRANSFERIDO" in movement:
                    meta["warning"] = "needs_review"
                income_kind = "Rendimento" if movement.startswith("RENDIMENTO") else "Dividend"
                logger.info("Imported income: %s", context)
                postings = [
                    data.Posting(cash_account, amount.Amount(abs(value), "BRL"), None, None, None, None),
                    data.Posting(self._income_account(institution, income_kind), amount.Amount(-abs(value), "BRL"), None, None, None, None),
                ]
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} {income_kind}", postings, **meta))
                continue

            if movement in {"JUROS SOBRE CAPITAL PROPRIO", "JUROS SOBRE CAPITAL PROPRIO - TRANSFERIDO"}:
                if value is None:
                    raise ValueError(f"{context}: missing value for JCP")
                if "TRANSFERIDO" in movement:
                    meta["warning"] = "needs_review"
                logger.info("Imported JCP: %s", context)
                postings = [
                    data.Posting(cash_account, amount.Amount(abs(value), "BRL"), None, None, None, None),
                    data.Posting(self._income_account(institution, "JCP"), amount.Amount(-abs(value), "BRL"), None, None, None, None),
                ]
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} JCP", postings, **meta))
                continue

            if movement == "PAGAMENTO DE JUROS":
                if value is None:
                    raise ValueError(f"{context}: missing value for interest payment")
                logger.info("Imported interest payment: %s", context)
                postings = [
                    data.Posting(cash_account, amount.Amount(abs(value), "BRL"), None, None, None, None),
                    data.Posting(self._income_account(institution, "Interest"), amount.Amount(-abs(value), "BRL"), None, None, None, None),
                ]
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Interest", postings, **meta))
                continue

            if movement in {"BONIFICACAO EM ATIVOS", "DESDOBRO", "GRUPAMENTO"}:
                if quantity is None:
                    raise ValueError(f"{context}: missing quantity for corporate action")
                if quantity == 0:
                    logger.info("Ignored zero-quantity corporate action: %s", context)
                    continue
                logger.warning(
                    "Imported split placeholder with ratio=1; update manually: %s (%s %s)",
                    context, movement, ticker,
                )
                entries.append(
                    data.Custom(
                        self._meta(filepath, lineno, **meta),
                        txn_date,
                        "split",
                        [self._custom_value(ticker), self._custom_value(Decimal("1"))],
                    )
                )
                continue

            if movement == "LEILAO DE FRACAO":
                if quantity is None or value is None:
                    raise ValueError(f"{context}: missing quantity/value for fraction auction")
                if quantity == 0:
                    raise ValueError(f"{context}: zero quantity for fraction auction")
                effective_unit_price = abs(value) / abs(quantity) if unit_price is None else unit_price
                gross_value = abs(quantity) * effective_unit_price
                fee_value = gross_value - abs(value)
                logger.info("Imported fraction auction: %s", context)
                postings = [
                    data.Posting(cash_account, amount.Amount(abs(value), "BRL"), None, None, None, None),
                    # FIX #1: use _empty_costspec() so Beancount can match and
                    # reduce an existing lot. None would create a new zero-cost
                    # position instead of reducing the one already in the books.
                    data.Posting(asset_account, amount.Amount(-abs(quantity), ticker), self._empty_costspec(), amount.Amount(effective_unit_price, "BRL"), None, None),
                ]
                if fee_value != 0:
                    postings.append(
                        data.Posting(
                            self._expense_account(institution, "Fees"),
                            amount.Amount(fee_value, "BRL"),
                            None, None, None, None,
                        )
                    )
                postings.append(
                    data.Posting(self._income_account(institution, "Gains"), None, None, None, None, None)
                )
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Fraction Auction", postings, **meta))
                continue

            if movement == "TRANSFERENCIA":
                if quantity is None:
                    raise ValueError(f"{context}: missing quantity for transfer")
                logger.info("Imported asset transfer: %s", context)
                signed_quantity = -abs(quantity) if direction == "DEBITO" else abs(quantity)
                postings = [
                    data.Posting(asset_account, self._amount(signed_quantity, ticker), self._empty_costspec(), None, None, None),
                    data.Posting("Equity:Transfers", self._amount(-signed_quantity, ticker), self._empty_costspec(), None, None, None),
                ]
                meta["warning"] = "unmatched_transfer"
                entries.append(self._txn(filepath, lineno, txn_date, f"Asset Transfer {ticker}", postings, **meta))
                continue

            if movement in {"RESGATE ANTECIPADO/", "RESGATE ANTECIPADO"}:
                if quantity is None or value is None:
                    raise ValueError(f"{context}: missing quantity/value for redemption")
                if quantity == 0:
                    raise ValueError(f"{context}: zero quantity for redemption")
                effective_price = abs(value) / abs(quantity) if unit_price is None else unit_price
                gross_value = abs(quantity) * effective_price
                fee_value = gross_value - abs(value)
                logger.info("Imported redemption: %s", context)
                postings = [
                    data.Posting(cash_account, self._amount(abs(value), "BRL"), None, None, None, None),
                    # FIX #1: same as LEILAO DE FRACAO above.
                    data.Posting(asset_account, self._amount(-abs(quantity), ticker), self._empty_costspec(), self._amount(effective_price, "BRL"), None, None),
                ]
                if fee_value != 0:
                    postings.append(
                        data.Posting(
                            self._expense_account(institution, "Fees"),
                            self._amount(fee_value, "BRL"),
                            None, None, None, None,
                        )
                    )
                postings.append(
                    data.Posting(self._income_account(institution, "Gains"), None, None, None, None, None)
                )
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Redemption", postings, **meta))
                continue

            if movement in {"COMPRA", "APLICACAO"}:
                if quantity is None or value is None:
                    raise ValueError(f"{context}: missing quantity/value for purchase")
                effective_price = abs(value) / abs(quantity) if unit_price is None else unit_price
                gross_value = abs(quantity) * effective_price
                fee_value = abs(value) - gross_value
                logger.info("Imported fixed income purchase: %s", context)
                postings = [
                    data.Posting(asset_account, amount.Amount(quantity, ticker), self._cost(effective_price, txn_date), None, None, None),
                    data.Posting(cash_account, amount.Amount(-abs(value), "BRL"), None, None, None, None),
                ]
                if fee_value != 0:
                    postings.append(
                        data.Posting(
                            self._expense_account(institution, "Fees"),
                            amount.Amount(fee_value, "BRL"),
                            None, None, None, None,
                        )
                    )
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Purchase", postings, **meta))
                continue

            if movement in {"COMPRA / VENDA", "COMPRA/VENDA"}:
                if quantity is None or value is None:
                    raise ValueError(f"{context}: missing quantity/value for compra/venda")
                effective_price = abs(value) / abs(quantity) if unit_price is None else unit_price
                if direction == "DEBITO":
                    gross_value = abs(quantity) * effective_price
                    fee_value = gross_value - abs(value)
                    logger.info("Imported fixed income sale: %s", context)
                    postings = [
                        data.Posting(cash_account, amount.Amount(abs(value), "BRL"), None, None, None, None),
                        # FIX #1: same as LEILAO DE FRACAO above.
                        data.Posting(asset_account, amount.Amount(-abs(quantity), ticker), self._empty_costspec(), amount.Amount(effective_price, "BRL"), None, None),
                    ]
                    if fee_value != 0:
                        postings.append(
                            data.Posting(
                                self._expense_account(institution, "Fees"),
                                amount.Amount(fee_value, "BRL"),
                                None, None, None, None,
                            )
                        )
                    postings.append(
                        data.Posting(self._income_account(institution, "Gains"), None, None, None, None, None)
                    )
                    entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Sale", postings, **meta))
                else:
                    gross_value = abs(quantity) * effective_price
                    fee_value = abs(value) - gross_value
                    logger.info("Imported fixed income purchase: %s", context)
                    postings = [
                        data.Posting(asset_account, amount.Amount(abs(quantity), ticker), self._cost(effective_price, txn_date), None, None, None),
                        data.Posting(cash_account, amount.Amount(-abs(value), "BRL"), None, None, None, None),
                    ]
                    if fee_value != 0:
                        postings.append(
                            data.Posting(
                                self._expense_account(institution, "Fees"),
                                amount.Amount(fee_value, "BRL"),
                                None, None, None, None,
                            )
                        )
                    entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Purchase", postings, **meta))
                continue

            if "TAXA" in movement or "COBRANCA" in movement:
                if value is None:
                    raise ValueError(f"{context}: missing value for fee")
                logger.info("Imported fee: %s", context)
                postings = [
                    data.Posting(self._expense_account(institution, "Fees"), amount.Amount(abs(value), "BRL"), None, None, None, None),
                    data.Posting(cash_account, amount.Amount(-abs(value), "BRL"), None, None, None, None),
                ]
                entries.append(self._txn(filepath, lineno, txn_date, f"{ticker} Fee", postings, **meta))
                continue

            logger.error("Unsupported movement type at %s: %s", context, movement)
            raise ValueError(f"{context}: unsupported movement type {movement!r}")

        return entries